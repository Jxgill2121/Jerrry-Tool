# Validation tab - Set validation limits and check cycle data against specifications

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import List, Dict, Optional

import pandas as pd

from powertech_tools.utils.file_parser import load_table_allow_duplicate_headers, build_minmax_display_map
from powertech_tools.utils.helpers import safe_float
from powertech_tools.data.validator import validate_maxmin_file


def build_tab(parent, app):
    """
    Build the validation tab UI.

    Args:
        parent: Parent frame for this tab
        app: Main application instance (for storing state)
    """
    f = ttk.Frame(parent)
    f.pack(fill="both", expand=True, padx=20, pady=20)

    # Import theme for styling
    from powertech_tools.config.theme import PowertechTheme

    title_label = ttk.Label(f, text="Validation & Quality Control", style='Title.TLabel')
    title_label.pack(anchor="w", pady=(0, 5))

    desc_label = ttk.Label(
        f,
        text="Set validation limits and check cycle data against specifications",
        style='Subtitle.TLabel'
    )
    desc_label.pack(anchor="w", pady=(0, 20))

    # File selection
    card1 = ttk.LabelFrame(f, text="Step 1: Load Max/Min File", padding=20)
    card1.pack(fill="x", pady=(0, 12))

    app.val_infile = tk.StringVar(value="")
    btn_frame = ttk.Frame(card1)
    btn_frame.pack(fill="x")

    ttk.Button(
        btn_frame,
        text="📁 Choose Max/Min File",
        command=lambda: _val_choose_file(app),
        style='Action.TButton'
    ).pack(side="left")

    ttk.Label(btn_frame, textvariable=app.val_infile, style='Status.TLabel').pack(side="left", padx=15)

    # Cycle column
    card2 = ttk.LabelFrame(f, text="Step 2: Select Cycle Column", padding=20)
    card2.pack(fill="x", pady=(0, 12))

    app.val_cycle_col = tk.StringVar(value="")
    app.cb_val_cycle = ttk.Combobox(card2, state="disabled", width=40, textvariable=app.val_cycle_col, values=[])
    app.cb_val_cycle.pack(side="left", padx=10)

    # Validation limits
    card3 = ttk.LabelFrame(f, text="Step 3: Set Validation Limits", padding=20)
    card3.pack(fill="both", expand=True, pady=(0, 12))

    ttk.Label(
        card3,
        text="Configure min/max acceptable ranges for each variable:",
        style='Subtitle.TLabel'
    ).pack(anchor="w", pady=(0, 10))

    canvas = tk.Canvas(card3, height=200, bg="white")
    scrollbar = ttk.Scrollbar(card3, orient="vertical", command=canvas.yview)
    app.val_limits_frame = ttk.Frame(canvas)

    app.val_limits_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=app.val_limits_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    hdr = ttk.Frame(app.val_limits_frame)
    hdr.pack(fill="x", pady=8)
    ttk.Label(hdr, text="Variable", width=18, font=(PowertechTheme.FONT_FAMILY, 9, 'bold')).grid(row=0, column=0, padx=5)
    ttk.Label(hdr, text="Min Lower", width=10, font=(PowertechTheme.FONT_FAMILY, 9, 'bold')).grid(row=0, column=1, padx=5)
    ttk.Label(hdr, text="Min Upper", width=10, font=(PowertechTheme.FONT_FAMILY, 9, 'bold')).grid(row=0, column=2, padx=5)
    ttk.Label(hdr, text="Max Lower", width=10, font=(PowertechTheme.FONT_FAMILY, 9, 'bold')).grid(row=0, column=3, padx=5)
    ttk.Label(hdr, text="Max Upper", width=10, font=(PowertechTheme.FONT_FAMILY, 9, 'bold')).grid(row=0, column=4, padx=5)

    app.val_limit_entries = []
    app.val_df = None
    app.val_variables = []

    # Actions
    card4 = ttk.LabelFrame(f, text="Step 4: Validation Actions", padding=20)
    card4.pack(fill="x", pady=(0, 12))

    action_frame = ttk.Frame(card4)
    action_frame.pack(fill="x")

    ttk.Button(
        action_frame,
        text="▶ RUN VALIDATION",
        command=lambda: _val_validate(app),
        style='Action.TButton'
    ).pack(side="left", padx=(0, 10))

    ttk.Button(
        action_frame,
        text="📊 EXPORT TO EXCEL",
        command=lambda: _val_export_excel(app)
    ).pack(side="left", padx=(0, 10))

    ttk.Button(
        action_frame,
        text="📥 LOAD LIMITS",
        command=lambda: _val_load_limits(app)
    ).pack(side="left", padx=(0, 10))

    ttk.Button(
        action_frame,
        text="💾 SAVE LIMITS",
        command=lambda: _val_save_limits(app)
    ).pack(side="left")

    app.val_status = tk.StringVar(value="")
    ttk.Label(action_frame, textvariable=app.val_status, style='Status.TLabel').pack(side="left", padx=15)

    # Results
    card5 = ttk.LabelFrame(f, text="Validation Results", padding=15)
    card5.pack(fill="both", expand=True)

    app.val_results_text = tk.Text(
        card5,
        height=12,
        wrap="none",
        bg="white",
        fg=PowertechTheme.TEXT_PRIMARY,
        font=(PowertechTheme.FONT_FAMILY, 9)
    )
    app.val_results_text.pack(fill="both", expand=True)

    app.val_results_df = None
    app.val_summary = None


def _val_choose_file(app):
    """Handle file selection for validation"""
    path = filedialog.askopenfilename(
        title="Select max/min file",
        filetypes=[("Text/Log", "*.txt *.log *.dat *.csv *.tsv"), ("All", "*.*")]
    )
    if not path:
        return
    app.val_infile.set(path)
    _val_load_file(app)


def _val_load_file(app):
    """Load and parse the validation file"""
    try:
        path = app.val_infile.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", f"File not found: {path}")
            return

        raw_df = load_table_allow_duplicate_headers(path)

        internal, internal_to_display, internal_kind = build_minmax_display_map(list(raw_df.columns))
        raw_df.columns = internal

        df = raw_df.rename(columns=internal_to_display)

        app.val_df = df
        headers = list(df.columns)

        variables = set()
        for h in headers:
            if h.endswith(" (Min)"):
                base = h[:-6]
                if f"{base} (Max)" in headers:
                    variables.add(base)

        app.val_variables = sorted(list(variables))

        app.cb_val_cycle["values"] = headers
        app.cb_val_cycle["state"] = "readonly"
        app.val_cycle_col.set("Cycle" if "Cycle" in headers else (headers[1] if len(headers) > 1 else headers[0]))

        for widget in app.val_limits_frame.winfo_children()[1:]:
            widget.destroy()
        app.val_limit_entries.clear()

        for var in app.val_variables:
            row = ttk.Frame(app.val_limits_frame)
            row.pack(fill="x", pady=3)

            ttk.Label(row, text=var, width=18).grid(row=0, column=0, padx=5, sticky="w")

            min_lower = tk.StringVar(value="")
            min_upper = tk.StringVar(value="")
            max_lower = tk.StringVar(value="")
            max_upper = tk.StringVar(value="")

            ttk.Entry(row, textvariable=min_lower, width=10).grid(row=0, column=1, padx=5)
            ttk.Entry(row, textvariable=min_upper, width=10).grid(row=0, column=2, padx=5)
            ttk.Entry(row, textvariable=max_lower, width=10).grid(row=0, column=3, padx=5)
            ttk.Entry(row, textvariable=max_upper, width=10).grid(row=0, column=4, padx=5)

            app.val_limit_entries.append({
                "variable": var,
                "min_lower": min_lower,
                "min_upper": min_upper,
                "max_lower": max_lower,
                "max_upper": max_upper
            })

        app.val_status.set(f"✓ Loaded {len(df)} cycles, {len(app.val_variables)} variables")

    except Exception as e:
        app.val_df = None
        messagebox.showerror("Error", str(e))


def _val_validate(app):
    """Run validation checks"""
    try:
        if app.val_df is None:
            messagebox.showerror("Error", "Please load a file first")
            return

        cycle_col = app.val_cycle_col.get().strip()
        if not cycle_col or cycle_col not in app.val_df.columns:
            messagebox.showerror("Error", "Invalid cycle column")
            return

        limits = {}
        for entry in app.val_limit_entries:
            var = entry["variable"]

            min_lower = safe_float(entry["min_lower"].get())
            min_upper = safe_float(entry["min_upper"].get())
            max_lower = safe_float(entry["max_lower"].get())
            max_upper = safe_float(entry["max_upper"].get())

            if min_lower == "INVALID" or min_upper == "INVALID" or max_lower == "INVALID" or max_upper == "INVALID":
                messagebox.showerror("Error", f"Invalid limits for '{var}'")
                return

            if any(x is not None for x in [min_lower, min_upper, max_lower, max_upper]):
                limits[var] = {
                    "min_lower": min_lower,
                    "min_upper": min_upper,
                    "max_lower": max_lower,
                    "max_upper": max_upper
                }

        if not limits:
            messagebox.showwarning("Warning", "Please set at least one limit")
            return

        results_df, summary = validate_maxmin_file(app.val_df, limits, cycle_col)
        app.val_results_df = results_df
        app.val_summary = summary

        app.val_results_text.delete("1.0", tk.END)

        summary_text = (
            "═══════════════════════════════════\n"
            "     VALIDATION SUMMARY\n"
            "═══════════════════════════════════\n"
            f"Total Cycles:  {summary['total_cycles']}\n"
            f"✓ Passed:      {summary['passed_cycles']} ({summary['pass_rate']:.1f}%)\n"
            f"✗ Failed:      {summary['failed_cycles']} ({100 - summary['pass_rate']:.1f}%)\n\n"
            "Violations by Variable:\n"
        )
        for var, count in sorted(summary["violation_by_variable"].items()):
            summary_text += f"  • {var}: {count} violation(s)\n"

        summary_text += "\n═══════════════════════════════════\n"
        summary_text += "  FAILED CYCLES (First 50)\n"
        summary_text += "═══════════════════════════════════\n\n"

        app.val_results_text.insert(tk.END, summary_text)
        failed = results_df[results_df["Status"] == "FAIL"].head(50)
        app.val_results_text.insert(tk.END, failed.to_string(index=False))

        app.val_status.set(f"✓ Validation complete: {summary['passed_cycles']}/{summary['total_cycles']} passed")
        messagebox.showinfo(
            "Validation Complete",
            f"✓ Passed: {summary['passed_cycles']}\n✗ Failed: {summary['failed_cycles']}\n\nExport to Excel for full results"
        )

    except Exception as e:
        messagebox.showerror("Error", str(e))


def _val_export_excel(app):
    """Export validation results to Excel"""
    try:
        if app.val_results_df is None or app.val_summary is None or app.val_df is None:
            messagebox.showerror("Error", "Please run validation first")
            return

        out_path = filedialog.asksaveasfilename(
            title="Save validation results",
            defaultextension=".xlsx",
            initialfile="validation_results.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")]
        )
        if not out_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            messagebox.showerror("Missing Library", "openpyxl required\n\nInstall: pip install openpyxl")
            return

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_data = [
            ["POWERTECH VALIDATION SUMMARY"],
            [""],
            ["Total Cycles", app.val_summary["total_cycles"]],
            ["Passed Cycles", app.val_summary["passed_cycles"]],
            ["Failed Cycles", app.val_summary["failed_cycles"]],
            ["Pass Rate (%)", f"{app.val_summary['pass_rate']:.1f}%"],
            [""],
            ["VIOLATIONS BY VARIABLE"],
        ]
        for var, count in sorted(app.val_summary["violation_by_variable"].items()):
            summary_data.append([var, count])

        for row_data in summary_data:
            ws_summary.append(row_data)

        title_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        ws_summary["A1"].fill = title_fill
        ws_summary["A1"].font = title_font

        ws_results = wb.create_sheet("Results")

        merged_df = app.val_df.copy()
        merged_df["Validation_Status"] = app.val_results_df["Status"]
        merged_df["Violations"] = app.val_results_df["Violations"]

        headers = ["Validation_Status", "Violations"] + list(app.val_df.columns)
        ws_results.append(headers)

        for idx, _row in merged_df.iterrows():
            status = app.val_results_df.loc[idx, "Status"]
            violations = app.val_results_df.loc[idx, "Violations"]
            row_data = [status, violations] + list(app.val_df.iloc[idx])
            ws_results.append(row_data)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws_results[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row_idx in range(2, ws_results.max_row + 1):
            status_cell = ws_results.cell(row=row_idx, column=1)
            fill = green_fill if status_cell.value == "PASS" else red_fill
            for col_idx in range(1, ws_results.max_column + 1):
                ws_results.cell(row=row_idx, column=col_idx).fill = fill

        for col in ws_results.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws_results.column_dimensions[col_letter].width = min(max_len + 2, 55)

        wb.save(out_path)
        messagebox.showinfo("Success", f"Excel file saved:\n{out_path}\n\n✓ Green = PASS\n✗ Red = FAIL")

    except Exception as e:
        messagebox.showerror("Error", str(e))


def _val_load_limits(app):
    """Load validation limits from CSV"""
    try:
        path = filedialog.askopenfilename(
            title="Load validation limits",
            filetypes=[("CSV", "*.csv"), ("All", "*.*")]
        )
        if not path:
            return

        limits_df = pd.read_csv(path)

        for entry in app.val_limit_entries:
            var = entry["variable"]
            row = limits_df[limits_df["Variable"] == var]
            if not row.empty:
                entry["min_lower"].set(str(row.iloc[0].get("Min_Lower", "")))
                entry["min_upper"].set(str(row.iloc[0].get("Min_Upper", "")))
                entry["max_lower"].set(str(row.iloc[0].get("Max_Lower", "")))
                entry["max_upper"].set(str(row.iloc[0].get("Max_Upper", "")))

        messagebox.showinfo("Success", f"Limits loaded from:\n{path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))


def _val_save_limits(app):
    """Save validation limits to CSV"""
    try:
        if not app.val_limit_entries:
            messagebox.showerror("Error", "Please load a file first")
            return

        out_path = filedialog.asksaveasfilename(
            title="Save validation limits",
            defaultextension=".csv",
            initialfile="validation_limits.csv",
            filetypes=[("CSV", "*.csv"), ("All", "*.*")]
        )
        if not out_path:
            return

        rows = []
        for entry in app.val_limit_entries:
            rows.append({
                "Variable": entry["variable"],
                "Min_Lower": entry["min_lower"].get(),
                "Min_Upper": entry["min_upper"].get(),
                "Max_Lower": entry["max_lower"].get(),
                "Max_Upper": entry["max_upper"].get()
            })

        pd.DataFrame(rows).to_csv(out_path, index=False)
        messagebox.showinfo("Success", f"Limits saved to:\n{out_path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))
