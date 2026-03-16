# ASR (Accelerated Stress Rupture) Validation tab
# Validates tank temperature hold duration for ASR testing

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import List, Optional

import pandas as pd

from powertech_tools.utils.helpers import safe_float, ScrollableFrame
from powertech_tools.data.asr_validator import (
    validate_asr_temperature,
    format_duration,
    load_asr_data_from_file
)


def build_tab(parent, app):
    """
    Build the ASR validation tab UI.

    Args:
        parent: Parent frame for this tab
        app: Main application instance (for storing state)
    """
    scrollable = ScrollableFrame(parent)
    scrollable.pack(fill="both", expand=True)
    f = ttk.Frame(scrollable.content)
    f.pack(fill="both", expand=True, padx=15, pady=15)

    # Import theme for styling
    from powertech_tools.config.theme import PowertechTheme

    title_label = ttk.Label(f, text="ASR Temperature Validation", style='Title.TLabel')
    title_label.pack(anchor="w", pady=(0, 5))

    desc_label = ttk.Label(
        f,
        text="Accelerated Stress Rupture - Calculate cumulative time within temperature band",
        style='Subtitle.TLabel'
    )
    desc_label.pack(anchor="w", pady=(0, 20))

    # File selection
    card1 = ttk.LabelFrame(f, text="Step 1: Load Test Data", padding=20)
    card1.pack(fill="x", pady=(0, 12))

    app.asr_infile = tk.StringVar(value="")
    btn_frame = ttk.Frame(card1)
    btn_frame.pack(fill="x")

    ttk.Button(
        btn_frame,
        text="Load TXT/CSV File",
        command=lambda: _asr_choose_txt(app),
        style='Action.TButton'
    ).pack(side="left", padx=(0, 10))

    ttk.Button(
        btn_frame,
        text="Load TDMS File",
        command=lambda: _asr_choose_tdms(app),
    ).pack(side="left")

    file_label_frame = ttk.Frame(card1)
    file_label_frame.pack(fill="x", pady=(10, 0))
    ttk.Label(file_label_frame, textvariable=app.asr_infile, style='Status.TLabel').pack(side="left")

    # Column selection
    card2 = ttk.LabelFrame(f, text="Step 2: Select Data Columns", padding=20)
    card2.pack(fill="x", pady=(0, 12))

    col_frame = ttk.Frame(card2)
    col_frame.pack(fill="x")

    # Time column
    time_row = ttk.Frame(col_frame)
    time_row.pack(fill="x", pady=5)
    ttk.Label(time_row, text="Time Column:", width=20).pack(side="left")
    app.asr_time_col = tk.StringVar(value="")
    app.cb_asr_time = ttk.Combobox(time_row, state="disabled", width=35, textvariable=app.asr_time_col, values=[])
    app.cb_asr_time.pack(side="left", padx=10)

    # Temperature column
    temp_row = ttk.Frame(col_frame)
    temp_row.pack(fill="x", pady=5)
    ttk.Label(temp_row, text="Temperature Column:", width=20).pack(side="left")
    app.asr_temp_col = tk.StringVar(value="")
    app.cb_asr_temp = ttk.Combobox(temp_row, state="disabled", width=35, textvariable=app.asr_temp_col, values=[])
    app.cb_asr_temp.pack(side="left", padx=10)

    # Temperature band configuration
    card3 = ttk.LabelFrame(f, text="Step 3: Set Temperature Band", padding=20)
    card3.pack(fill="x", pady=(0, 12))

    band_frame = ttk.Frame(card3)
    band_frame.pack(fill="x")

    ttk.Label(
        card3,
        text="Set the acceptable temperature range. Time within this band counts toward ASR duration.",
        style='Subtitle.TLabel'
    ).pack(anchor="w", pady=(0, 10))

    # Min temp
    min_row = ttk.Frame(band_frame)
    min_row.pack(fill="x", pady=5)
    ttk.Label(min_row, text="Minimum Temperature:", width=20).pack(side="left")
    app.asr_temp_min = tk.StringVar(value="83")
    ttk.Entry(min_row, textvariable=app.asr_temp_min, width=15).pack(side="left", padx=10)
    ttk.Label(min_row, text="C").pack(side="left")

    # Max temp
    max_row = ttk.Frame(band_frame)
    max_row.pack(fill="x", pady=5)
    ttk.Label(max_row, text="Maximum Temperature:", width=20).pack(side="left")
    app.asr_temp_max = tk.StringVar(value="87")
    ttk.Entry(max_row, textvariable=app.asr_temp_max, width=15).pack(side="left", padx=10)
    ttk.Label(max_row, text="C").pack(side="left")

    # Target duration (optional reference)
    target_row = ttk.Frame(band_frame)
    target_row.pack(fill="x", pady=5)
    ttk.Label(target_row, text="Target Duration (optional):", width=20).pack(side="left")
    app.asr_target_hours = tk.StringVar(value="22")
    ttk.Entry(target_row, textvariable=app.asr_target_hours, width=15).pack(side="left", padx=10)
    ttk.Label(target_row, text="hours").pack(side="left")

    # Actions
    card4 = ttk.LabelFrame(f, text="Step 4: Run Validation", padding=20)
    card4.pack(fill="x", pady=(0, 12))

    action_frame = ttk.Frame(card4)
    action_frame.pack(fill="x")

    ttk.Button(
        action_frame,
        text="CALCULATE ASR DURATION",
        command=lambda: _asr_validate(app),
        style='Action.TButton'
    ).pack(side="left", padx=(0, 10))

    ttk.Button(
        action_frame,
        text="EXPORT TO EXCEL",
        command=lambda: _asr_export_excel(app)
    ).pack(side="left", padx=(0, 10))

    app.asr_status = tk.StringVar(value="")
    ttk.Label(action_frame, textvariable=app.asr_status, style='Status.TLabel').pack(side="left", padx=15)

    # Results
    card5 = ttk.LabelFrame(f, text="ASR Validation Results", padding=15)
    card5.pack(fill="both", expand=True)

    app.asr_results_text = tk.Text(
        card5,
        height=15,
        wrap="none",
        bg="white",
        fg=PowertechTheme.TEXT_PRIMARY,
        font=(PowertechTheme.FONT_FAMILY, 10)
    )
    app.asr_results_text.pack(fill="both", expand=True)

    # Scrollbars for results
    xscroll = ttk.Scrollbar(card5, orient="horizontal", command=app.asr_results_text.xview)
    xscroll.pack(fill="x")
    app.asr_results_text.configure(xscrollcommand=xscroll.set)

    # Store data
    app.asr_df = None
    app.asr_summary = None
    app.asr_detail_df = None


def _asr_choose_txt(app):
    """Load TXT/CSV file for ASR validation"""
    path = filedialog.askopenfilename(
        title="Select ASR test data file",
        filetypes=[("Text/CSV", "*.txt *.csv *.log *.dat *.tsv"), ("All", "*.*")]
    )
    if not path:
        return

    try:
        df, columns = load_asr_data_from_file(path)
        app.asr_df = df
        app.asr_infile.set(path)

        # Update column dropdowns
        app.cb_asr_time["values"] = columns
        app.cb_asr_time["state"] = "readonly"
        app.cb_asr_temp["values"] = columns
        app.cb_asr_temp["state"] = "readonly"

        # Auto-select time column
        time_candidates = ['time', 'Time', 'TIME', 't', 'T', 'Elapsed Time', 'Elapsed_Time']
        for tc in time_candidates:
            if tc in columns:
                app.asr_time_col.set(tc)
                break
        else:
            if columns:
                app.asr_time_col.set(columns[0])

        # Auto-select temperature column (look for common temp names)
        temp_candidates = ['twall', 'Twall', 'TWALL', 'Temperature', 'Temp', 'temp', 'T_wall']
        for tc in temp_candidates:
            for col in columns:
                if tc.lower() in col.lower():
                    app.asr_temp_col.set(col)
                    break
            if app.asr_temp_col.get():
                break

        app.asr_status.set(f"Loaded {len(df)} data points, {len(columns)} columns")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file: {e}")


def _asr_choose_tdms(app):
    """Load TDMS file for ASR validation"""
    path = filedialog.askopenfilename(
        title="Select TDMS file",
        filetypes=[("TDMS", "*.tdms"), ("All", "*.*")]
    )
    if not path:
        return

    try:
        from nptdms import TdmsFile

        tdms_file = TdmsFile.read(path)

        # Get all groups and channels
        all_channels = []
        channel_data = {}

        for group in tdms_file.groups():
            for channel in group.channels():
                full_name = f"{group.name}/{channel.name}"
                all_channels.append(full_name)

                # Get data
                data = channel[:]
                if hasattr(data, '__len__') and len(data) > 0:
                    channel_data[full_name] = data

                    # Try to get time track
                    try:
                        time_track = channel.time_track()
                        if time_track is not None and len(time_track) == len(data):
                            channel_data[f"{full_name}_time"] = time_track
                    except Exception:
                        pass

        # Build DataFrame from all channels with same length
        if channel_data:
            # Find most common length
            lengths = [len(v) for v in channel_data.values()]
            common_len = max(set(lengths), key=lengths.count)

            # Filter to channels with common length
            filtered_data = {k: v for k, v in channel_data.items() if len(v) == common_len}

            df = pd.DataFrame(filtered_data)

            # If no time column, generate one
            if 'Time' not in df.columns:
                # Look for any time track
                time_cols = [c for c in df.columns if '_time' in c.lower()]
                if time_cols:
                    df['Time'] = df[time_cols[0]]
                else:
                    # Generate synthetic time (0.1s intervals)
                    df.insert(0, 'Time', range(len(df)) * 0.1)

            app.asr_df = df
            columns = list(df.columns)

            app.asr_infile.set(path)
            app.cb_asr_time["values"] = columns
            app.cb_asr_time["state"] = "readonly"
            app.cb_asr_temp["values"] = columns
            app.cb_asr_temp["state"] = "readonly"

            # Auto-select time
            if 'Time' in columns:
                app.asr_time_col.set('Time')
            elif columns:
                app.asr_time_col.set(columns[0])

            # Auto-select temp
            temp_candidates = ['twall', 'Twall', 'TWALL', 'Temperature', 'Temp']
            for tc in temp_candidates:
                for col in columns:
                    if tc.lower() in col.lower():
                        app.asr_temp_col.set(col)
                        break
                if app.asr_temp_col.get():
                    break

            app.asr_status.set(f"Loaded {len(df)} data points from TDMS")

    except ImportError:
        messagebox.showerror("Missing Library", "nptdms required for TDMS files\n\nInstall: pip install nptdms")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load TDMS: {e}")


def _asr_validate(app):
    """Run ASR temperature validation"""
    try:
        if app.asr_df is None:
            messagebox.showerror("Error", "Please load a data file first")
            return

        time_col = app.asr_time_col.get().strip()
        temp_col = app.asr_temp_col.get().strip()

        if not time_col or time_col not in app.asr_df.columns:
            messagebox.showerror("Error", "Please select a valid time column")
            return

        if not temp_col or temp_col not in app.asr_df.columns:
            messagebox.showerror("Error", "Please select a valid temperature column")
            return

        # Parse temperature band
        temp_min = safe_float(app.asr_temp_min.get())
        temp_max = safe_float(app.asr_temp_max.get())

        if temp_min == "INVALID" or temp_min is None:
            messagebox.showerror("Error", "Invalid minimum temperature")
            return
        if temp_max == "INVALID" or temp_max is None:
            messagebox.showerror("Error", "Invalid maximum temperature")
            return
        if temp_min >= temp_max:
            messagebox.showerror("Error", "Minimum must be less than maximum")
            return

        # Parse target hours (optional)
        target_hours = safe_float(app.asr_target_hours.get())
        if target_hours == "INVALID":
            target_hours = None

        # Run validation
        summary, detail_df = validate_asr_temperature(
            app.asr_df,
            time_col,
            temp_col,
            temp_min,
            temp_max,
            time_unit="seconds"  # Keep in seconds for precision
        )

        app.asr_summary = summary
        app.asr_detail_df = detail_df

        # Display results
        app.asr_results_text.delete("1.0", tk.END)

        # Format durations
        total_sec = summary['total_duration']
        in_band_sec = summary['time_in_band']
        out_band_sec = summary['time_out_band']

        total_fmt = format_duration(total_sec, "auto")
        in_band_fmt = format_duration(in_band_sec, "auto")
        out_band_fmt = format_duration(out_band_sec, "auto")

        # Hours for target comparison
        in_band_hours = in_band_sec / 3600

        results_text = (
            "=" * 60 + "\n"
            "        ASR TEMPERATURE VALIDATION RESULTS\n"
            "=" * 60 + "\n\n"
            f"Temperature Band: {temp_min}C - {temp_max}C\n"
            f"Data Points: {summary['data_points']:,}\n\n"
            "-" * 40 + "\n"
            "          DURATION SUMMARY\n"
            "-" * 40 + "\n\n"
            f"  Total Test Duration:    {total_fmt}\n"
            f"  Time IN BAND:           {in_band_fmt}  ({summary['percent_in_band']:.2f}%)\n"
            f"  Time OUT OF BAND:       {out_band_fmt}  ({summary['percent_out_band']:.2f}%)\n\n"
            f"  Time in band (hours):   {in_band_hours:.4f} hrs\n\n"
        )

        # Target comparison if provided
        if target_hours is not None and target_hours > 0:
            progress = (in_band_hours / target_hours) * 100
            status = "TARGET MET" if in_band_hours >= target_hours else "TARGET NOT MET"
            remaining = max(0, target_hours - in_band_hours)

            results_text += (
                "-" * 40 + "\n"
                "          TARGET COMPARISON\n"
                "-" * 40 + "\n\n"
                f"  Target Duration:        {target_hours:.2f} hours\n"
                f"  Achieved:               {in_band_hours:.4f} hours\n"
                f"  Progress:               {progress:.1f}%\n"
                f"  Status:                 {status}\n"
            )
            if remaining > 0:
                results_text += f"  Remaining:              {remaining:.4f} hours\n"
            results_text += "\n"

        # Temperature statistics
        stats = summary['temp_stats']
        results_text += (
            "-" * 40 + "\n"
            "          TEMPERATURE STATISTICS\n"
            "-" * 40 + "\n\n"
            f"  Minimum:    {stats['min']:.2f}C\n"
            f"  Maximum:    {stats['max']:.2f}C\n"
            f"  Mean:       {stats['mean']:.2f}C\n"
            f"  Std Dev:    {stats['std']:.2f}C\n\n"
        )

        # Excursion summary
        results_text += (
            "-" * 40 + "\n"
            "          EXCURSIONS (Out of Band)\n"
            "-" * 40 + "\n\n"
            f"  Total Excursions: {summary['excursion_count']}\n\n"
        )

        # List first 20 excursions
        excursions = summary['excursions']
        if excursions:
            results_text += "  Time (s)        Duration      Temp Range\n"
            results_text += "  " + "-" * 50 + "\n"
            for i, exc in enumerate(excursions[:20]):
                start = exc['start_time']
                dur = format_duration(exc['duration'], "auto")
                min_t = exc['min_temp']
                max_t = exc['max_temp']
                temp_range = f"{min_t:.1f} - {max_t:.1f}C" if min_t and max_t else "N/A"
                results_text += f"  {start:>10.1f}     {dur:<12}  {temp_range}\n"

            if len(excursions) > 20:
                results_text += f"\n  ... and {len(excursions) - 20} more excursions\n"
        else:
            results_text += "  No excursions - temperature stayed within band entire test!\n"

        results_text += "\n" + "=" * 60 + "\n"

        app.asr_results_text.insert(tk.END, results_text)
        app.asr_status.set(f"Validation complete: {in_band_fmt} in band ({summary['percent_in_band']:.1f}%)")

    except Exception as e:
        messagebox.showerror("Error", str(e))
        import traceback
        traceback.print_exc()


def _asr_export_excel(app):
    """Export ASR results to Excel"""
    try:
        if app.asr_summary is None or app.asr_detail_df is None:
            messagebox.showerror("Error", "Please run validation first")
            return

        out_path = filedialog.asksaveasfilename(
            title="Save ASR results",
            defaultextension=".xlsx",
            initialfile="asr_validation_results.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")]
        )
        if not out_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            messagebox.showerror("Missing Library", "openpyxl required\n\nInstall: pip install openpyxl")
            return

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary = app.asr_summary
        in_band_hours = summary['time_in_band'] / 3600
        total_hours = summary['total_duration'] / 3600

        summary_data = [
            ["ASR VALIDATION SUMMARY"],
            [""],
            ["Temperature Band", f"{summary['temp_band'][0]}C - {summary['temp_band'][1]}C"],
            ["Data Points", summary['data_points']],
            [""],
            ["DURATION RESULTS"],
            ["Total Duration (hours)", f"{total_hours:.4f}"],
            ["Time In Band (hours)", f"{in_band_hours:.4f}"],
            ["Time Out of Band (hours)", f"{(summary['time_out_band'] / 3600):.4f}"],
            ["Percent In Band", f"{summary['percent_in_band']:.2f}%"],
            [""],
            ["TEMPERATURE STATISTICS"],
            ["Minimum", f"{summary['temp_stats']['min']:.2f}C"],
            ["Maximum", f"{summary['temp_stats']['max']:.2f}C"],
            ["Mean", f"{summary['temp_stats']['mean']:.2f}C"],
            ["Std Dev", f"{summary['temp_stats']['std']:.2f}C"],
            [""],
            ["Excursion Count", summary['excursion_count']],
        ]

        for row_data in summary_data:
            ws_summary.append(row_data)

        # Style header
        title_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        ws_summary["A1"].fill = title_fill
        ws_summary["A1"].font = title_font

        # Excursions sheet
        ws_exc = wb.create_sheet("Excursions")
        ws_exc.append(["Start Time (s)", "End Time (s)", "Duration (s)", "Min Temp (C)", "Max Temp (C)"])

        for exc in summary['excursions']:
            ws_exc.append([
                exc['start_time'],
                exc['end_time'],
                exc['duration'],
                exc['min_temp'],
                exc['max_temp']
            ])

        # Style excursions header
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_exc[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Detail data sheet
        ws_detail = wb.create_sheet("Detail Data")
        detail_cols = list(app.asr_detail_df.columns)
        ws_detail.append(detail_cols)

        # Add detail data (limit to first 100k rows for Excel performance)
        for idx, row in app.asr_detail_df.head(100000).iterrows():
            ws_detail.append(list(row))

        # Style detail header
        for cell in ws_detail[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Conditional formatting for In_Band column
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Find Status column index
        try:
            status_col_idx = detail_cols.index('Status') + 1
            for row_idx in range(2, min(ws_detail.max_row + 1, 100002)):
                cell = ws_detail.cell(row=row_idx, column=status_col_idx)
                if cell.value == 'IN BAND':
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
        except (ValueError, IndexError):
            pass

        wb.save(out_path)
        messagebox.showinfo("Success", f"Excel file saved:\n{out_path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))
