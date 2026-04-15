"""
Validation & Quality Control tab
Check cycle max/min data against configurable limits per variable
"""

import io
import os
import tempfile

import pandas as pd
import streamlit as st

from powertech_tools.data.validator import validate_maxmin_file
from powertech_tools.utils.file_parser import build_minmax_display_map, load_table_allow_duplicate_headers
from powertech_tools.utils.helpers import safe_float

_ACCEPTED_TYPES = ["txt", "log", "dat", "csv", "tsv"]


def _load_val_file(uploaded_file) -> tuple:
    """Save uploaded file to temp dir, load and return (df, variables, headers)."""
    tmpdir = tempfile.mkdtemp()
    path = os.path.join(tmpdir, uploaded_file.name)
    with open(path, "wb") as fh:
        fh.write(uploaded_file.getbuffer())

    raw_df = load_table_allow_duplicate_headers(path)
    internal, internal_to_display, _ = build_minmax_display_map(list(raw_df.columns))
    raw_df.columns = internal
    df = raw_df.rename(columns=internal_to_display)

    headers = list(df.columns)
    variables = sorted({
        h[:-6] for h in headers
        if h.endswith(" (Min)") and f"{h[:-6]} (Max)" in headers
    })
    return df, variables, headers


def render():
    st.subheader("Validation & Quality Control")
    st.caption("Set validation limits and check cycle data against specifications")

    # --- Step 1: Load max/min file ---
    st.markdown("### Step 1: Load Max/Min File")
    uploaded_file = st.file_uploader(
        "Choose Max/Min File",
        type=_ACCEPTED_TYPES,
        accept_multiple_files=False,
        key="val_uploaded_file",
    )

    if not uploaded_file:
        st.info("Upload a max/min file (output from the Max/Min Converter tab).")
        return

    if st.session_state.get("val_file_name") != uploaded_file.name:
        try:
            df, variables, headers = _load_val_file(uploaded_file)
        except Exception as e:
            st.error(f"Could not load file: {e}")
            return

        st.session_state["val_file_name"] = uploaded_file.name
        st.session_state["val_df"] = df
        st.session_state["val_variables"] = variables
        st.session_state["val_headers"] = headers
        # Reset limits table with blank entries
        st.session_state["val_limits_df"] = pd.DataFrame(
            [{"Variable": v, "Min Lower": None, "Min Upper": None,
              "Max Lower": None, "Max Upper": None}
             for v in variables]
        )
        st.session_state.pop("val_results_df", None)
        st.session_state.pop("val_summary", None)

    df: pd.DataFrame = st.session_state["val_df"]
    variables: list = st.session_state["val_variables"]
    headers: list = st.session_state["val_headers"]

    st.success(f"Loaded {len(df)} cycle(s) — {len(variables)} variable pair(s)")

    # --- Step 2: Cycle column ---
    st.markdown("### Step 2: Select Cycle Column")
    default_cycle = "Cycle" if "Cycle" in headers else (headers[1] if len(headers) > 1 else headers[0])
    cycle_col = st.selectbox(
        "Cycle Column",
        headers,
        index=headers.index(default_cycle) if default_cycle in headers else 0,
        key="val_cycle_col",
    )

    # --- Step 3: Validation limits ---
    st.markdown("### Step 3: Set Validation Limits")
    st.caption(
        "Leave any field blank to skip that check. "
        "Min Lower/Upper = acceptable range for the minimum value per cycle. "
        "Max Lower/Upper = acceptable range for the maximum value per cycle."
    )

    # Optional: load limits from CSV
    with st.expander("Load limits from CSV"):
        limits_upload = st.file_uploader(
            "Upload limits CSV",
            type=["csv"],
            key="val_limits_upload",
        )
        if limits_upload is not None:
            try:
                loaded = pd.read_csv(limits_upload)
                existing = st.session_state["val_limits_df"].copy()
                for _, row in loaded.iterrows():
                    mask = existing["Variable"] == row.get("Variable", "")
                    if mask.any():
                        for col in ["Min Lower", "Min Upper", "Max Lower", "Max Upper"]:
                            csv_key = col.replace(" ", "_")
                            if csv_key in row and pd.notna(row[csv_key]):
                                existing.loc[mask, col] = row[csv_key]
                st.session_state["val_limits_df"] = existing
                st.success("Limits loaded — scroll down to review.")
                st.rerun()
            except Exception as e:
                st.error(f"Could not load limits CSV: {e}")

    limits_df: pd.DataFrame = st.session_state["val_limits_df"]

    edited_limits = st.data_editor(
        limits_df,
        column_config={
            "Variable": st.column_config.TextColumn("Variable", disabled=True),
            "Min Lower": st.column_config.NumberColumn("Min Lower", format="%.4f"),
            "Min Upper": st.column_config.NumberColumn("Min Upper", format="%.4f"),
            "Max Lower": st.column_config.NumberColumn("Max Lower", format="%.4f"),
            "Max Upper": st.column_config.NumberColumn("Max Upper", format="%.4f"),
        },
        use_container_width=True,
        hide_index=True,
        key="val_limits_editor",
        num_rows="fixed",
    )

    # Save limits download
    limits_csv = io.StringIO()
    col_rename = {"Min Lower": "Min_Lower", "Min Upper": "Min_Upper",
                  "Max Lower": "Max_Lower", "Max Upper": "Max_Upper"}
    edited_limits.rename(columns=col_rename).to_csv(limits_csv, index=False)
    st.download_button(
        label="Save Limits as CSV",
        data=limits_csv.getvalue().encode(),
        file_name="validation_limits.csv",
        mime="text/csv",
    )

    # --- Step 4: Run validation ---
    st.markdown("### Step 4: Run Validation")

    if st.button("RUN VALIDATION", type="primary", key="val_run_btn"):
        # Build limits dict from editor
        limits = {}
        for _, row in edited_limits.iterrows():
            var = row["Variable"]
            vals = {
                "min_lower": safe_float(str(row["Min Lower"])) if pd.notna(row["Min Lower"]) else None,
                "min_upper": safe_float(str(row["Min Upper"])) if pd.notna(row["Min Upper"]) else None,
                "max_lower": safe_float(str(row["Max Lower"])) if pd.notna(row["Max Lower"]) else None,
                "max_upper": safe_float(str(row["Max Upper"])) if pd.notna(row["Max Upper"]) else None,
            }
            # Only include if at least one limit is set
            if any(v is not None and v != "INVALID" for v in vals.values()):
                limits[var] = vals

        if not limits:
            st.error("Please set at least one limit in the table above.")
        elif not cycle_col or cycle_col not in df.columns:
            st.error("Please select a valid cycle column.")
        else:
            with st.spinner("Running validation..."):
                try:
                    results_df, summary = validate_maxmin_file(df, limits, cycle_col)
                    st.session_state["val_results_df"] = results_df
                    st.session_state["val_summary"] = summary
                    st.session_state["val_df_snap"] = df.copy()
                except Exception as e:
                    st.error(f"Validation failed: {e}")

    # --- Results ---
    if "val_results_df" not in st.session_state:
        return

    results_df: pd.DataFrame = st.session_state["val_results_df"]
    summary: dict = st.session_state["val_summary"]
    df_snap: pd.DataFrame = st.session_state["val_df_snap"]

    # Summary metrics
    st.markdown("### Results")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Cycles", summary["total_cycles"])
    c2.metric("Passed", summary["passed_cycles"])
    c3.metric("Failed", summary["failed_cycles"])
    c4.metric("Pass Rate", f"{summary['pass_rate']:.1f}%")

    # Violations by variable
    if summary["violation_by_variable"]:
        st.markdown("**Violations by Variable**")
        viol_df = pd.DataFrame(
            [{"Variable": k, "Violations": v}
             for k, v in sorted(summary["violation_by_variable"].items())]
        )
        st.dataframe(viol_df, use_container_width=True, hide_index=True)

    # Failed cycles table
    failed = results_df[results_df["Status"] == "FAIL"]
    if not failed.empty:
        st.markdown(f"**Failed Cycles ({len(failed)})**")
        st.dataframe(failed[["Cycle", "Status", "Violations"]], use_container_width=True, hide_index=True)
    else:
        st.success("All cycles passed!")

    # Full results table
    with st.expander("Full Results Table"):
        st.dataframe(results_df[["Cycle", "Status", "Violations"]], use_container_width=True, hide_index=True)

    # Excel export
    st.markdown("### Export")
    if st.button("Generate Excel Report", key="val_excel_btn"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws_sum = wb.active
            ws_sum.title = "Summary"

            title_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
            title_font = Font(bold=True, color="FFFFFF", size=14)

            summary_rows = [
                ["POWERTECH VALIDATION SUMMARY"], [""],
                ["Total Cycles", summary["total_cycles"]],
                ["Passed Cycles", summary["passed_cycles"]],
                ["Failed Cycles", summary["failed_cycles"]],
                ["Pass Rate (%)", f"{summary['pass_rate']:.1f}%"],
                [""], ["VIOLATIONS BY VARIABLE"],
            ]
            for var, count in sorted(summary["violation_by_variable"].items()):
                summary_rows.append([var, count])
            for r in summary_rows:
                ws_sum.append(r)
            ws_sum["A1"].fill = title_fill
            ws_sum["A1"].font = title_font

            # Results sheet with colour coding
            ws_res = wb.create_sheet("Results")
            merged = df_snap.copy()
            merged.insert(0, "Validation_Status", results_df["Status"].values)
            merged.insert(1, "Violations", results_df["Violations"].values)

            ws_res.append(list(merged.columns))
            hdr_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")
            for cell in ws_res[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font

            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for row_vals in merged.itertuples(index=False):
                ws_res.append(list(row_vals))
            for row_idx in range(2, ws_res.max_row + 1):
                fill = green if ws_res.cell(row_idx, 1).value == "PASS" else red
                for col_idx in range(1, ws_res.max_column + 1):
                    ws_res.cell(row_idx, col_idx).fill = fill

            # Auto-width
            for col in ws_res.columns:
                letter = col[0].column_letter
                width = max((len(str(c.value or "")) for c in col), default=8)
                ws_res.column_dimensions[letter].width = min(width + 2, 55)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                label="Download Excel Report (.xlsx)",
                data=buf,
                file_name="validation_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except ImportError:
            st.error("openpyxl is required. Run: pip install openpyxl")
        except Exception as e:
            st.error(f"Excel export failed: {e}")
