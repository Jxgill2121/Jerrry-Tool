"""
ASR Validation tab
Accelerated Stress Rupture - cumulative time within value bands per parameter
"""

import io
import os
import tempfile

import pandas as pd
import streamlit as st

from powertech_tools.data.asr_validator import (
    format_duration,
    load_asr_data_from_file,
    validate_asr_temperature,
)
from powertech_tools.utils.helpers import natural_sort_key, safe_float

_ACCEPTED_TYPES = ["txt", "csv", "log", "dat", "tsv"]

# Default band values keyed by substring of column name (lower-case)
_DEFAULT_BANDS = {
    "tskin":       (83, 87),
    "twall":       (83, 87),
    "tfluid":      (83, 87),
    "tfuel":       (83, 87),
    "temp":        (83, 87),
    "temperature": (83, 87),
    "ptank":       (0, 100),
    "pressure":    (0, 100),
}

_EXCLUDE_COLS = {"time", "elapsed", "t", "_file_index", "_file_name", "datetime", "date"}


def _band_defaults(param: str):
    p = param.lower()
    for key, (lo, hi) in _DEFAULT_BANDS.items():
        if key in p:
            return lo, hi
    return 83, 87


def render():
    st.subheader("ASR Validation")
    st.caption(
        "Accelerated Stress Rupture — Calculate cumulative time within value bands "
        "for multiple parameters"
    )

    # --- Step 1: Load files ---
    st.markdown("### Step 1: Load Test Data")
    uploaded_files = st.file_uploader(
        "Choose TXT/CSV Files",
        type=_ACCEPTED_TYPES,
        accept_multiple_files=True,
        key="asr_uploaded_files",
    )

    if not uploaded_files:
        st.info("Upload one or more test data files to get started.")
        return

    uploaded_files = sorted(uploaded_files, key=lambda f: natural_sort_key(f.name))
    current_names = [f.name for f in uploaded_files]

    if st.session_state.get("asr_file_names") != current_names:
        tmpdir = tempfile.mkdtemp()
        saved_paths = []
        for uf in uploaded_files:
            path = os.path.join(tmpdir, uf.name)
            with open(path, "wb") as fh:
                fh.write(uf.getbuffer())
            saved_paths.append(path)

        try:
            all_dfs = []
            columns = None
            for i, path in enumerate(saved_paths):
                df, cols = load_asr_data_from_file(path)
                df["_file_index"] = i
                df["_file_name"] = os.path.basename(path)
                all_dfs.append(df)
                if columns is None:
                    columns = cols
            combined = pd.concat(all_dfs, ignore_index=True)
        except Exception as e:
            st.error(f"Failed to load files: {e}")
            return

        st.session_state["asr_file_names"] = current_names
        st.session_state["asr_saved_paths"] = saved_paths
        st.session_state["asr_df"] = combined
        st.session_state["asr_columns"] = columns
        st.session_state.pop("asr_results", None)

    combined: pd.DataFrame = st.session_state["asr_df"]
    columns: list = st.session_state["asr_columns"]

    st.success(
        f"{len(uploaded_files)} file(s) — "
        f"{len(combined):,} data points — "
        f"{len(columns)} columns"
    )

    # --- Step 2: Time column + target ---
    st.markdown("### Step 2: Time Configuration")
    col1, col2 = st.columns(2)

    with col1:
        time_candidates = ["time", "Time", "TIME", "Elapsed", "elapsed", "t"]
        def _time_idx():
            for c in time_candidates:
                if c in columns:
                    return columns.index(c)
            return 0
        time_col = st.selectbox("Time Column", columns, index=_time_idx(), key="asr_time_col")

    with col2:
        target_hours = st.number_input(
            "Target Duration (hours, optional — 0 to skip)",
            min_value=0.0,
            value=22.0,
            step=1.0,
            key="asr_target_hours",
        )

    # --- Step 3: Parameter bands ---
    st.markdown("### Step 3: Select Parameters & Set Bands")
    st.caption("Check parameters to validate and set the min/max band for each.")

    param_columns = [
        c for c in columns
        if c.lower() not in _EXCLUDE_COLS and not c.startswith("_")
    ]

    if not param_columns:
        st.error("No parameter columns found.")
        return

    # Build the editable band table
    bands_init = pd.DataFrame(
        [
            {
                "Parameter": p,
                "Selected": False,
                "Min": float(_band_defaults(p)[0]),
                "Max": float(_band_defaults(p)[1]),
            }
            for p in param_columns
        ]
    )

    edited_bands = st.data_editor(
        bands_init,
        column_config={
            "Selected": st.column_config.CheckboxColumn("Select", default=False),
            "Min": st.column_config.NumberColumn("Min", format="%.2f"),
            "Max": st.column_config.NumberColumn("Max", format="%.2f"),
        },
        use_container_width=True,
        hide_index=True,
        key="asr_bands_editor",
        num_rows="fixed",
    )

    # --- Step 4: Run validation ---
    st.markdown("### Step 4: Run Validation")

    if st.button("CALCULATE ASR DURATION", type="primary", key="asr_run_btn"):
        selected_rows = edited_bands[edited_bands["Selected"]]
        if selected_rows.empty:
            st.error("Please select at least one parameter in the table above.")
        elif not time_col or time_col not in combined.columns:
            st.error("Please select a valid time column.")
        else:
            with st.spinner("Running validation..."):
                results = {}
                errors = []

                # Detect date range for display
                date_range_str = None
                dt_cols = [c for c in combined.columns if "datetime" in c.lower() or "date" in c.lower()]
                if dt_cols:
                    try:
                        dt_vals = pd.to_datetime(combined[dt_cols[0]], errors="coerce").dropna()
                        if not dt_vals.empty:
                            date_range_str = (
                                f"{dt_vals.min().strftime('%Y-%m-%d %H:%M')} "
                                f"to {dt_vals.max().strftime('%Y-%m-%d %H:%M')}"
                            )
                    except Exception:
                        pass

                for _, row in selected_rows.iterrows():
                    param = row["Parameter"]
                    v_min, v_max = row["Min"], row["Max"]

                    if v_min >= v_max:
                        errors.append(f"{param}: Min must be less than Max.")
                        continue
                    if param not in combined.columns:
                        errors.append(f"{param}: column not found in data.")
                        continue

                    try:
                        summary, detail_df = validate_asr_temperature(
                            combined, time_col, param, v_min, v_max, time_unit="seconds"
                        )
                        results[param] = {
                            "summary": summary,
                            "detail_df": detail_df,
                            "band": (v_min, v_max),
                        }
                    except Exception as e:
                        errors.append(f"{param}: {e}")

                if errors:
                    for err in errors:
                        st.warning(err)

                st.session_state["asr_results"] = results
                st.session_state["asr_date_range"] = date_range_str
                st.session_state["asr_target_hours_snap"] = target_hours

    # --- Results ---
    if "asr_results" not in st.session_state:
        return

    results: dict = st.session_state["asr_results"]
    date_range_str: str = st.session_state.get("asr_date_range")
    target_h: float = st.session_state.get("asr_target_hours_snap", 0.0)

    if not results:
        st.info("No results — check that the selected parameters exist in the data.")
        return

    st.success(f"Validated {len(results)} parameter(s)")

    if date_range_str:
        st.caption(f"Date range: {date_range_str}")

    # Summary table
    st.markdown("### Results Summary")
    summary_rows = []
    for param, data in results.items():
        s = data["summary"]
        band = data["band"]
        in_band_h = s["time_in_band"] / 3600
        total_h = s["total_duration"] / 3600
        target_met = ""
        if target_h and target_h > 0:
            target_met = "Yes" if in_band_h >= target_h else "No"
        summary_rows.append(
            {
                "Parameter": param,
                "Band": f"{band[0]} – {band[1]}",
                "Total (h)": f"{total_h:.4f}",
                "In Band (h)": f"{in_band_h:.4f}",
                "In Band %": f"{s['percent_in_band']:.2f}%",
                "Mean": f"{s['temp_stats']['mean']:.2f}",
                "Excursions": s["excursion_count"],
                **({"Target Met": target_met} if target_met != "" else {}),
            }
        )
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    # Per-parameter detail cards
    st.markdown("### Per-Parameter Detail")
    for param, data in results.items():
        s = data["summary"]
        band = data["band"]
        in_band_h = s["time_in_band"] / 3600

        with st.expander(f"{param}  —  band {band[0]}–{band[1]}"):
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total Duration", format_duration(s["total_duration"], "auto"))
            c2.metric("Time In Band", format_duration(s["time_in_band"], "auto"))
            c3.metric("In Band %", f"{s['percent_in_band']:.2f}%")
            c4.metric("Excursions", s["excursion_count"])

            c1b, c2b, c3b = st.columns(3)
            c1b.metric("Min", f"{s['temp_stats']['min']:.2f}")
            c2b.metric("Mean", f"{s['temp_stats']['mean']:.2f}")
            c3b.metric("Max", f"{s['temp_stats']['max']:.2f}")

            if target_h and target_h > 0:
                progress_pct = min(in_band_h / target_h, 1.0)
                met = in_band_h >= target_h
                st.markdown(
                    f"**Target:** {target_h:.2f} h — "
                    f"{'Met' if met else 'Not Met'} "
                    f"({(in_band_h / target_h * 100):.1f}%)"
                )
                st.progress(progress_pct)

    # Excel export
    st.markdown("### Export")
    if st.button("Generate Excel Report", key="asr_excel_btn"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"

            hdr_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")

            target_h_snap = st.session_state.get("asr_target_hours_snap", 0.0)
            headers = [
                "Parameter", "Band Min", "Band Max", "Total (hrs)", "In Band (hrs)",
                "In Band %", "Min Value", "Max Value", "Mean", "Excursions", "Target Met",
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font

            for param, data in results.items():
                s = data["summary"]
                band = data["band"]
                in_band_h = s["time_in_band"] / 3600
                total_h = s["total_duration"] / 3600
                stats = s["temp_stats"]
                target_met = ""
                if target_h_snap and target_h_snap > 0:
                    target_met = "Yes" if in_band_h >= target_h_snap else "No"
                ws.append([
                    param, band[0], band[1],
                    f"{total_h:.4f}", f"{in_band_h:.4f}",
                    f"{s['percent_in_band']:.2f}%",
                    f"{stats['min']:.2f}", f"{stats['max']:.2f}", f"{stats['mean']:.2f}",
                    s["excursion_count"], target_met,
                ])

            for param, data in results.items():
                sheet_name = param[:28].replace("/", "_").replace("\\", "_").replace("*", "_")
                ws_d = wb.create_sheet(f"{sheet_name}_Exc")
                ws_d.append(["Start Time (s)", "End Time (s)", "Duration (s)", "Min Value", "Max Value"])
                for cell in ws_d[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                for exc in data["summary"]["excursions"]:
                    ws_d.append([
                        exc["start_time"], exc["end_time"], exc["duration"],
                        exc["min_temp"], exc["max_temp"],
                    ])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                label="Download Excel Report (.xlsx)",
                data=buf,
                file_name="asr_validation_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except ImportError:
            st.error("openpyxl is required for Excel export. Run: pip install openpyxl")
        except Exception as e:
            st.error(f"Excel export failed: {e}")
